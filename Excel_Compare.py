import os
import json
import sys
from openpyxl import load_workbook

# 設定命令行輸出格式
def print_step(step, message):
    print(f"\n{'='*70}")
    print(f"【{step}】{message}")
    print(f"{'='*70}")

def print_sub_step(message):
    print(f"  ▶ {message}")

def print_warning(message):
    print(f"\n{'!'*70}")
    print(f"【警告】{message}")
    print(f"{'!'*70}")

def print_error(message):
    print(f"\n{'#'*70}")
    print(f"【錯誤】{message}")
    print(f"{'#'*70}")

def print_info(title, message):
    print(f"\n{'-'*60}")
    print(f"【{title}】{message}")
    print(f"{'-'*60}")

# 讀取或建立設定檔 (JSON)
CONFIG_FILE = 'config_compare.json'
def load_or_create_config():
    print_info("設定檔", "讀取或建立設定檔...")
    default_config = {
        'Data_1': 'Data_1.xlsx',  # 比較的 Excel-1
        'Data_2': 'Data_2.xlsx',  # 比較的 Excel-2
        'Data_1_Sheet': 'Data',  # Excel-1 的分頁
        'Data_2_Sheet': 'Data', # Excel-2 的分頁
        'OUTPUT_FILE': 'Compare_Result.txt', # 對比差異的輸出紀錄
        'Scan_Method': 0  # 是否使用逐行掃描 (較慢但詳細)
    }
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                config = json.load(f)
                is_modified = False
                for key, value in default_config.items():
                    if key not in config:
                        config[key] = value
                        is_modified = True
                        print_sub_step(f"新增預設設定: {key} = {value}")
                
                if is_modified:
                    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                        json.dump(config, f, ensure_ascii=False, indent=4)
                    print_sub_step("已更新設定檔")
                else:
                    print_sub_step("成功讀取現有設定檔")
        else:
            config = default_config
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=4)
            print_sub_step("建立新的設定檔")
        return config
    except Exception as e:
        print_error(f"設定檔操作失敗 - {str(e)}")
        sys.exit(1)

# 讀取 Excel 檔
def load_excel(file_path):
    try:
        return load_workbook(file_path)
    except FileNotFoundError:
        print_error(f"找不到 Excel 檔案: {file_path}")
    except Exception as e:
        print_error(f"無法讀取 Excel 檔案 {file_path}，錯誤: {e}")
    sys.exit(1)

# 主程序
def main():
    # 載入設定
    config = load_or_create_config()
    OUTPUT_FILE = config['OUTPUT_FILE']
    SCAN_METHOD = bool(config.get('Scan_Method', 1))  # 轉換為布林值，確保 1/0 仍可運作

    # 讀取兩個 Excel 文件
    wb1 = load_excel(config['Data_1'])
    wb2 = load_excel(config['Data_2'])

    # 讀取工作表
    sheet1 = wb1[config['Data_1_Sheet']]
    sheet2 = wb2[config['Data_2_Sheet']]

    # 用於儲存差異的清單
    differences = []

    if SCAN_METHOD:
        # 使用 iter_rows() 逐行遍歷
        print_info("比對模式", "使用 iter_rows() 逐行遍歷")
        for row_num, row in enumerate(sheet1.iter_rows(), 1):
            for col_num, cell in enumerate(row, 1):
                cell1_value = cell.value
                cell2_value = sheet2.cell(row=row_num, column=col_num).value
                if cell1_value != cell2_value:
                    differences.append((row_num, col_num, cell1_value, cell2_value))
    else:
        # 使用最大範圍掃描
        print_info("比對模式", "使用最大範圍掃描")
        max_rows = max(sheet1.max_row, sheet2.max_row)
        max_cols = max(sheet1.max_column, sheet2.max_column)
        for row_num in range(1, max_rows + 1):
            for col_num in range(1, max_cols + 1):
                cell1_value = sheet1.cell(row=row_num, column=col_num).value
                cell2_value = sheet2.cell(row=row_num, column=col_num).value
                if cell1_value != cell2_value:
                    differences.append((row_num, col_num, cell1_value, cell2_value))

    # 輸出結果
    if differences:
        with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
            f.write(f"發現 {len(differences)} 個差異\n")
            for row, col, val1, val2 in differences:
                f.write(f"[行 {row}, 列 {col}] -> {val1} ≠ {val2}\n")
        print_info("比對結果", f"發現 {len(differences)} 個差異，已輸出至 {OUTPUT_FILE}")
    else:
        print_step("比對結果", "兩個文件完全相同！")


if __name__ == "__main__":
    main()
